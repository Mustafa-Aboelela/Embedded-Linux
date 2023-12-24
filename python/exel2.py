from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook('hello_world.xlsx')  # Replace 'example.xlsx' with your Excel file name

# Select the active sheet
sheet = workbook.active

# Access data in a specific cell
cell_value = sheet['A1'].value  # Reads the value of cell A1
print("Value in cell A1:", cell_value)

# Iterating through rows and columns to read data
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):  # Read first 5 rows and first 3 columns
    for cell in row:
        print(cell.value, end="\t")  # Print cell value
    print()  # Move to the next line for the next row

# Close the workbook
workbook.close()
